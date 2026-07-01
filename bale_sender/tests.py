from pathlib import Path
from io import BytesIO
import tempfile
from unittest.mock import Mock, patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from .core import BaleSafirClient, ExcelRecipient, build_excel_preview, normalize_iran_mobile, process_excel_batch, read_excel_recipients, render_message, run_excel_batch, send_single_recipient_test
from .forms import SingleMessageTestForm, UploadExcelForm
from .models import MessageBatch, MessageRecipient


TEST_PHONE_LOCAL = "09120000000"
TEST_PHONE_STD = "989120000000"


class PhoneTests(TestCase):
    def test_normalize_iran_mobile(self):
        self.assertEqual(normalize_iran_mobile(TEST_PHONE_LOCAL), TEST_PHONE_STD)
        self.assertEqual(normalize_iran_mobile("+" + TEST_PHONE_STD), TEST_PHONE_STD)
        self.assertEqual(normalize_iran_mobile("۹۱۲۰۰۰۰۰۰۰"), TEST_PHONE_STD)
        self.assertIsNone(normalize_iran_mobile("123"))


class ExcelTests(TestCase):
    def make_excel(self):
        tmp = Path(tempfile.mkdtemp()) / "sample.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["ردیف", "نام", "نام خانوادگی", "موبایل"])
        ws.append([1, "علی", "رضایی", TEST_PHONE_LOCAL])
        ws.append([2, "زهرا", "احمدی", "123"])
        wb.save(tmp)
        return tmp

    def test_read_recipients(self):
        rows = read_excel_recipients(self.make_excel())
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].full_name, "علی رضایی")
        self.assertEqual(rows[0].normalized_phone, TEST_PHONE_STD)
        self.assertIsNone(rows[1].normalized_phone)

    def test_read_recipients_invalid_sheet_message(self):
        with self.assertRaisesMessage(ValueError, "شیت «missing» در فایل اکسل پیدا نشد"):
            read_excel_recipients(self.make_excel(), sheet_name="missing")

    def test_read_recipients_missing_required_column_message(self):
        tmp = Path(tempfile.mkdtemp()) / "missing.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["نام", "موبایل"])
        wb.save(tmp)

        with self.assertRaisesMessage(ValueError, "ستون «نام خانوادگی»"):
            read_excel_recipients(tmp)

    def test_read_recipients_empty_header_message(self):
        tmp = Path(tempfile.mkdtemp()) / "empty_header.xlsx"
        wb = Workbook()
        wb.active.append([None, None, None])
        wb.save(tmp)

        with self.assertRaisesMessage(ValueError, "ردیف اول فایل اکسل باید header"):
            read_excel_recipients(tmp)

    def test_build_excel_preview_counts_and_limits_rows(self):
        tmp = Path(tempfile.mkdtemp()) / "preview.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["نام", "نام خانوادگی", "موبایل"])
        ws.append(["علی", "رضایی", TEST_PHONE_LOCAL])
        ws.append(["تکراری", "رضایی", TEST_PHONE_LOCAL])
        ws.append(["بد", "شماره", "123"])
        wb.save(tmp)

        preview = build_excel_preview(tmp, message_template="سلام {full_name}", limit=2)
        self.assertEqual(preview["total_rows"], 3)
        self.assertEqual(preview["valid_rows"], 1)
        self.assertEqual(preview["duplicate_rows"], 1)
        self.assertEqual(preview["invalid_rows"], 1)
        self.assertEqual(len(preview["rows"]), 2)
        self.assertEqual(preview["rows"][0]["status"], "ok")

    def test_build_excel_preview_applies_data_range(self):
        preview = build_excel_preview(
            self.make_excel(),
            message_template="سلام {full_name}",
            range_start=2,
            range_end=2,
        )

        self.assertEqual(preview["total_file_rows"], 2)
        self.assertEqual(preview["total_rows"], 1)
        self.assertEqual(preview["rows"][0]["row_number"], 3)
        self.assertEqual(preview["rows"][0]["first_name"], "زهرا")


class MessageTemplateTests(TestCase):
    def test_render_message_allows_only_known_placeholders(self):
        recipient = ExcelRecipient(2, "علی", "رضایی", TEST_PHONE_LOCAL, TEST_PHONE_STD)
        self.assertEqual(render_message("سلام {full_name} {phone}", recipient), f"سلام علی رضایی {TEST_PHONE_STD}")

        with self.assertRaisesMessage(ValueError, "placeholderهای نامعتبر"):
            render_message("سلام {unknown}", recipient)

    def test_form_rejects_invalid_template_and_large_file(self):
        upload = SimpleUploadedFile("sample.xlsx", b"x" * 6)
        with override_settings(BALE_MAX_UPLOAD_SIZE_MB=0):
            form = UploadExcelForm(data={"message_template": "سلام {unknown}", "send_mode": "dry_run"}, files={"excel_file": upload})
            self.assertFalse(form.is_valid())
            self.assertIn("message_template", form.errors)
            self.assertIn("excel_file", form.errors)

    def test_upload_form_accepts_saved_preview_token_without_new_file(self):
        form = UploadExcelForm(
            data={
                "uploaded_file_token": "token",
                "message_template": "سلام {full_name}",
                "send_mode": "dry_run",
                "sleep_seconds": "0",
            }
        )
        self.assertTrue(form.is_valid())

    def test_upload_form_rejects_reversed_data_range(self):
        form = UploadExcelForm(
            data={
                "uploaded_file_token": "token",
                "message_template": "سلام {full_name}",
                "send_mode": "dry_run",
                "sleep_seconds": "0",
                "range_start": "101",
                "range_end": "100",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("range_end", form.errors)

    def test_single_message_test_form_validates_phone_and_template(self):
        form = SingleMessageTestForm(
            data={
                "first_name": "علی",
                "last_name": "رضایی",
                "phone": TEST_PHONE_LOCAL,
                "message_template": "سلام {full_name}",
            }
        )
        self.assertTrue(form.is_valid())

        bad_form = SingleMessageTestForm(
            data={
                "first_name": "علی",
                "last_name": "رضایی",
                "phone": "123",
                "message_template": "سلام {unknown}",
            }
        )
        self.assertFalse(bad_form.is_valid())
        self.assertIn("phone", bad_form.errors)
        self.assertIn("message_template", bad_form.errors)


class SafirClientTests(TestCase):
    @override_settings(BALE_API_ACCESS_KEY="test-key", BALE_BOT_ID=123456, BALE_SEND_URL="https://example.test/api", BALE_REQUEST_TIMEOUT=5)
    @patch("bale_sender.core.requests.post")
    def test_send_success(self, post_mock):
        resp = Mock()
        resp.status_code = 200
        resp.text = "{}"
        resp.json.return_value = {"ok": True}
        post_mock.return_value = resp
        result = BaleSafirClient().send(phone_number=TEST_PHONE_STD, text="سلام", request_id="r1")
        self.assertEqual(result["status"], MessageRecipient.Status.SENT)

    @override_settings(BALE_API_ACCESS_KEY="test-key", BALE_BOT_ID=123456, BALE_SEND_URL="https://example.test/api", BALE_REQUEST_TIMEOUT=5)
    @patch("bale_sender.core.requests.post")
    def test_not_bale_user(self, post_mock):
        resp = Mock()
        resp.status_code = 400
        resp.text = "{}"
        resp.json.return_value = {"code": "NotBaleUser", "message": "user not found"}
        post_mock.return_value = resp
        result = BaleSafirClient().send(phone_number=TEST_PHONE_STD, text="سلام", request_id="r1")
        self.assertEqual(result["status"], MessageRecipient.Status.NOT_BALE_USER)


class BatchTests(TestCase):
    def make_excel(self):
        tmp = Path(tempfile.mkdtemp()) / "sample.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["ردیف", "نام", "نام خانوادگی", "موبایل"])
        ws.append([1, "علی", "رضایی", TEST_PHONE_LOCAL])
        ws.append([2, "علی", "رضایی", TEST_PHONE_LOCAL])
        ws.append([3, "بد", "شماره", "123"])
        wb.save(tmp)
        return tmp

    def test_dry_run_batch_report(self):
        report = Path(tempfile.mkdtemp()) / "report.xlsx"
        batch = run_excel_batch(file_path=str(self.make_excel()), message_template="سلام {full_name}", dry_run=True, report_path=str(report))
        self.assertEqual(batch.total_rows, 3)
        self.assertEqual(batch.recipients.filter(status=MessageRecipient.Status.DRY_RUN).count(), 1)
        self.assertEqual(batch.recipients.filter(status=MessageRecipient.Status.DUPLICATE).count(), 1)
        self.assertEqual(batch.recipients.filter(status=MessageRecipient.Status.INVALID_PHONE).count(), 1)
        self.assertTrue(report.exists())

    def test_dry_run_batch_applies_data_range(self):
        report = Path(tempfile.mkdtemp()) / "range.xlsx"
        batch = run_excel_batch(
            file_path=str(self.make_excel()),
            message_template="سلام {full_name}",
            dry_run=True,
            report_path=str(report),
            range_start=2,
            range_end=3,
        )

        self.assertEqual(batch.range_start, 2)
        self.assertEqual(batch.range_end, 3)
        self.assertEqual(batch.total_rows, 2)
        self.assertEqual(list(batch.recipients.values_list("row_number", flat=True)), [3, 4])
        self.assertEqual(batch.recipients.filter(status=MessageRecipient.Status.DRY_RUN).count(), 1)
        self.assertEqual(batch.recipients.filter(status=MessageRecipient.Status.INVALID_PHONE).count(), 1)

    @patch("bale_sender.core.BaleSafirClient.send")
    def test_single_recipient_test_send_creates_reported_batch(self, send_mock):
        send_mock.return_value = {"status": MessageRecipient.Status.SENT, "http_status": 200, "api_message": "ok"}

        batch = send_single_recipient_test(
            first_name="علی",
            last_name="رضایی",
            phone=TEST_PHONE_LOCAL,
            message_template="سلام {full_name}",
        )

        self.assertEqual(batch.status, MessageBatch.Status.FINISHED)
        self.assertEqual(batch.total_rows, 1)
        self.assertEqual(batch.total_sent, 1)
        self.assertEqual(batch.recipients.get().final_text, "سلام علی رضایی")

    @patch("bale_sender.core.BaleSafirClient.send")
    def test_running_batch_can_be_cancelled_between_sends(self, send_mock):
        report = Path(tempfile.mkdtemp()) / "cancelled.xlsx"
        batch = MessageBatch.objects.create(
            source_file_name="sample.xlsx",
            message_template="سلام {full_name}",
            dry_run=False,
        )

        def request_cancel(**kwargs):
            MessageBatch.objects.filter(pk=batch.pk).update(cancel_requested=True)
            return {"status": MessageRecipient.Status.SENT, "http_status": 200, "api_message": "ok"}

        send_mock.side_effect = request_cancel

        process_excel_batch(
            batch=batch,
            file_path=str(self.make_excel()),
            sleep_seconds=0,
            report_path=str(report),
        )

        batch.refresh_from_db()
        self.assertEqual(batch.status, MessageBatch.Status.CANCELLED)
        self.assertTrue(batch.cancel_requested)
        self.assertEqual(batch.total_rows, 1)
        self.assertTrue(report.exists())


class ReportViewTests(TestCase):
    def make_batch_with_recipients(self, count=105):
        batch = MessageBatch.objects.create(source_file_name="sample.xlsx", message_template="سلام {full_name}")
        MessageRecipient.objects.bulk_create(
            [
                MessageRecipient(
                    batch=batch,
                    row_number=i,
                    first_name=f"نام {i}",
                    last_name="تست",
                    full_name=f"نام {i} تست",
                    raw_phone=TEST_PHONE_LOCAL,
                    normalized_phone=TEST_PHONE_STD,
                    status=MessageRecipient.Status.DRY_RUN,
                    final_text=f"سلام نام {i} تست",
                )
                for i in range(1, count + 1)
            ]
        )
        return batch

    def test_batch_detail_uses_paginated_recent_rows(self):
        batch = self.make_batch_with_recipients()

        response = self.client.get(reverse("bale_batch_detail", args=[batch.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_recipients"], 105)
        self.assertEqual(len(response.context["recipients"]), 100)
        self.assertEqual(response.context["recipients"][0].row_number, 105)


    def test_batch_live_status_returns_last_ten_rows(self):
        batch = self.make_batch_with_recipients(count=12)

        response = self.client.get(reverse("bale_batch_live_status", args=[batch.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["batch"]["id"], batch.id)
        self.assertEqual(len(payload["recent_recipients"]), 10)
        self.assertEqual(payload["recent_recipients"][0]["row_number"], 12)
        self.assertIn("status_label", payload["recent_recipients"][0])

    def test_recent_recipients_excel_report_limit(self):
        self.make_batch_with_recipients(count=3)

        response = self.client.get(reverse("bale_recent_recipients_report"), {"limit": 2})

        self.assertEqual(response.status_code, 200)
        self.assertIn("bale_recent_2_recipients.xlsx", response["Content-Disposition"])
        wb = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0][0], "شناسه گزارش")

    def test_dashboard_can_reuse_previous_batch_file_for_next_range(self):
        source = Path(tempfile.mkdtemp()) / "sample.xlsx"
        source.write_bytes(b"placeholder")
        batch = MessageBatch.objects.create(
            source_file_name=source.name,
            source_file_path=str(source),
            message_template="سلام {full_name}",
            range_start=1,
            range_end=100,
            total_rows=100,
        )

        response = self.client.get(reverse("bale_dashboard"), {"reuse_batch": batch.id})

        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertTrue(form.initial["uploaded_file_token"])
        self.assertEqual(form.initial["message_template"], "سلام {full_name}")
        self.assertEqual(form.initial["range_start"], 101)

    def test_cancel_batch_view_marks_running_batch_for_stop(self):
        batch = MessageBatch.objects.create(
            source_file_name="sample.xlsx",
            message_template="سلام",
            status=MessageBatch.Status.RUNNING,
        )

        with tempfile.TemporaryDirectory() as base_dir, override_settings(BASE_DIR=base_dir):
            response = self.client.post(reverse("bale_cancel_batch", args=[batch.id]))

        self.assertRedirects(response, reverse("bale_batch_detail", args=[batch.id]))
        batch.refresh_from_db()
        self.assertTrue(batch.cancel_requested)
        self.assertEqual(batch.status, MessageBatch.Status.CANCELLED)
        self.assertIsNotNone(batch.finished_at)
        self.assertIn("متوقف شد", batch.error_message)
        self.assertTrue(batch.report_path)

    def test_live_status_treats_cancel_requested_batch_as_inactive(self):
        batch = MessageBatch.objects.create(
            source_file_name="sample.xlsx",
            message_template="سلام",
            status=MessageBatch.Status.RUNNING,
            cancel_requested=True,
        )

        response = self.client.get(reverse("bale_batch_live_status", args=[batch.id]))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()["batch"]["is_active"])
