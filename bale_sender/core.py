from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from time import sleep
from uuid import uuid4
import json
import re
import string
from zipfile import BadZipFile

import requests
from django.conf import settings
from django.db.models import Count
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import MessageBatch, MessageRecipient

PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
ALLOWED_MESSAGE_PLACEHOLDERS = frozenset({"first_name", "last_name", "full_name", "phone"})


@dataclass
class ExcelRecipient:
    row_number: int
    first_name: str
    last_name: str
    raw_phone: str
    normalized_phone: str | None

    @property
    def full_name(self) -> str:
        return " ".join(x for x in [self.first_name, self.last_name] if x).strip()


def _recipient_preview_status(item: ExcelRecipient, seen: set[str]) -> tuple[str, str]:
    if not item.normalized_phone:
        return "invalid_phone", "شماره نامعتبر"
    if item.normalized_phone in seen:
        return "duplicate", "تکراری در فایل"
    seen.add(item.normalized_phone)
    return "ok", "آماده ارسال"


def normalize_iran_mobile(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip().translate(PERSIAN_DIGITS)
    s = re.sub(r"[^0-9+]", "", s)
    if s.startswith("+98"):
        s = "98" + s[3:]
    elif s.startswith("0098"):
        s = "98" + s[4:]
    elif s.startswith("09") and len(s) == 11:
        s = "98" + s[1:]
    elif s.startswith("9") and len(s) == 10:
        s = "98" + s
    if re.fullmatch(r"989\d{9}", s):
        return s
    return None


def _clean_header(value: object) -> str:
    return str(value or "").strip().replace("ي", "ی").replace("ك", "ک")


def _find_header(headers: list[str], candidates: set[str], title: str) -> int:
    for i, h in enumerate(headers):
        if h in candidates:
            return i
    visible_headers = "، ".join(h for h in headers if h) or "بدون عنوان"
    accepted_headers = "، ".join(sorted(candidates))
    raise ValueError(f"ستون «{title}» در فایل اکسل پیدا نشد. عنوان‌های قابل قبول برای این ستون: {accepted_headers}. ستون‌های موجود: {visible_headers}")


def validate_message_template(template: str) -> None:
    try:
        parsed = list(string.Formatter().parse(template))
    except ValueError as exc:
        raise ValueError("قالب پیام نامعتبر است. اگر می‌خواهید آکولاد معمولی نمایش دهید از {{ و }} استفاده کنید.") from exc

    invalid_fields: set[str] = set()
    for _, field_name, format_spec, conversion in parsed:
        if field_name is None:
            continue
        if format_spec or conversion:
            raise ValueError("قالب پیام فقط placeholder ساده را می‌پذیرد؛ از format spec یا conversion استفاده نکنید.")
        root_field = field_name.split(".", 1)[0].split("[", 1)[0]
        if not root_field:
            raise ValueError("قالب پیام نامعتبر است؛ placeholder خالی {} مجاز نیست.")
        if root_field != field_name or root_field not in ALLOWED_MESSAGE_PLACEHOLDERS:
            invalid_fields.add(field_name)

    if invalid_fields:
        allowed = "، ".join(f"{{{name}}}" for name in sorted(ALLOWED_MESSAGE_PLACEHOLDERS))
        invalid = "، ".join(f"{{{name}}}" for name in sorted(invalid_fields))
        raise ValueError(f"placeholderهای نامعتبر در متن پیام: {invalid}. فقط این موارد مجاز هستند: {allowed}.")


def read_excel_recipients(file_path: str | Path, sheet_name: str | None = None) -> list[ExcelRecipient]:
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise ValueError("فایل اکسل قابل خواندن نیست. لطفاً یک فایل سالم با پسوند xlsx یا xlsm بارگذاری کنید.") from exc

    try:
        ws = wb[sheet_name] if sheet_name else wb.active
    except KeyError as exc:
        available_sheets = "، ".join(wb.sheetnames)
        raise ValueError(f"شیت «{sheet_name}» در فایل اکسل پیدا نشد. شیت‌های موجود: {available_sheets}") from exc

    rows = ws.iter_rows(values_only=True)
    try:
        headers = [_clean_header(x) for x in next(rows)]
    except StopIteration as exc:
        raise ValueError("فایل اکسل خالی است و ردیف header ندارد.") from exc

    if not any(headers):
        raise ValueError("ردیف اول فایل اکسل باید header ستون‌ها باشد، اما خالی است.")

    first_idx = _find_header(headers, {"نام", "اسم", "first_name"}, "نام")
    last_idx = _find_header(headers, {"نام خانوادگی", "فامیلی", "last_name"}, "نام خانوادگی")
    phone_idx = _find_header(headers, {"موبایل", "شماره موبایل", "شماره همراه", "mobile", "phone"}, "موبایل")

    result: list[ExcelRecipient] = []
    for row_number, row in enumerate(rows, start=2):
        first_name = str(row[first_idx] or "").strip()
        last_name = str(row[last_idx] or "").strip()
        raw_phone = str(row[phone_idx] or "").strip()
        if not any([first_name, last_name, raw_phone]):
            continue
        result.append(
            ExcelRecipient(
                row_number=row_number,
                first_name=first_name,
                last_name=last_name,
                raw_phone=raw_phone,
                normalized_phone=normalize_iran_mobile(raw_phone),
            )
        )
    return result


def render_message(template: str, recipient: ExcelRecipient) -> str:
    validate_message_template(template)
    try:
        return template.format(
            first_name=recipient.first_name,
            last_name=recipient.last_name,
            full_name=recipient.full_name,
            phone=recipient.normalized_phone or recipient.raw_phone,
        )
    except (KeyError, ValueError) as exc:
        raise ValueError("قالب پیام نامعتبر است و امکان ساخت متن پیام وجود ندارد.") from exc


def build_excel_preview(file_path: str | Path, *, sheet_name: str | None = None, message_template: str, limit: int = 10) -> dict:
    recipients = read_excel_recipients(file_path, sheet_name=sheet_name)
    seen: set[str] = set()
    preview_rows: list[dict] = []
    counts = {"ok": 0, "invalid_phone": 0, "duplicate": 0}

    for item in recipients:
        status, status_label = _recipient_preview_status(item, seen)
        counts[status] += 1
        if len(preview_rows) < limit:
            preview_rows.append(
                {
                    "row_number": item.row_number,
                    "first_name": item.first_name,
                    "last_name": item.last_name,
                    "full_name": item.full_name,
                    "raw_phone": item.raw_phone,
                    "normalized_phone": item.normalized_phone or "",
                    "status": status,
                    "status_label": status_label,
                    "final_text": render_message(message_template, item),
                }
            )

    return {
        "total_rows": len(recipients),
        "valid_rows": counts["ok"],
        "invalid_rows": counts["invalid_phone"],
        "duplicate_rows": counts["duplicate"],
        "rows": preview_rows,
        "shown_rows": len(preview_rows),
        "is_sendable": counts["ok"] > 0,
    }


class BaleSafirClient:
    def __init__(self) -> None:
        self.url = settings.BALE_SEND_URL
        self.api_key = settings.BALE_API_ACCESS_KEY
        self.bot_id = settings.BALE_BOT_ID
        self.timeout = settings.BALE_REQUEST_TIMEOUT

    def send(self, *, phone_number: str, text: str, request_id: str, button_text: str | None = None, button_url: str | None = None) -> dict:
        if not self.api_key or not self.bot_id:
            return {"status": MessageRecipient.Status.CONFIG_ERROR, "error": "BALE_API_ACCESS_KEY یا BALE_BOT_ID تنظیم نشده است."}

        message: dict = {"text": text}
        if button_text and button_url:
            message["reply_markup"] = {"inline_keyboard": [[{"text": button_text, "url": button_url}]]}

        payload = {
            "request_id": request_id,
            "bot_id": self.bot_id,
            "phone_number": phone_number,
            "message_data": {"message": message},
        }
        headers = {"api-access-key": self.api_key, "Content-Type": "application/json"}

        try:
            response = requests.post(self.url, headers=headers, json=payload, timeout=self.timeout)
            try:
                data = response.json()
            except ValueError:
                data = {"raw": response.text}
        except requests.RequestException as exc:
            return {"status": MessageRecipient.Status.FAILED, "error": str(exc)}

        code = str(data.get("code") or data.get("error") or "")
        message_text = str(data.get("message") or data.get("description") or "")
        raw = json.dumps(data, ensure_ascii=False)

        if 200 <= response.status_code < 300:
            status = MessageRecipient.Status.SENT
        elif code == "NotBaleUser":
            status = MessageRecipient.Status.NOT_BALE_USER
        elif response.status_code == 429 or code == "RateLimitExceeded":
            status = MessageRecipient.Status.RATE_LIMITED
        elif response.status_code == 402 or code == "PaymentRequired":
            status = MessageRecipient.Status.PAYMENT_REQUIRED
        elif code in {"InvalidPhone", "InvalidPhoneNumber"}:
            status = MessageRecipient.Status.INVALID_PHONE
        else:
            status = MessageRecipient.Status.FAILED

        return {"status": status, "http_status": response.status_code, "api_code": code, "api_message": message_text, "raw_response": raw}


def write_report(batch: MessageBatch, report_path: Path) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش ارسال"
    ws.append(["ردیف اکسل", "نام", "نام خانوادگی", "نام کامل", "شماره خام", "شماره استاندارد", "وضعیت", "HTTP", "کد API", "پیام API", "خطا", "متن نهایی"])
    for r in batch.recipients.all().order_by("row_number", "id").iterator(chunk_size=1000):
        ws.append([r.row_number, r.first_name, r.last_name, r.full_name, r.raw_phone, r.normalized_phone, r.status, r.http_status, r.api_code, r.api_message, r.error_message, r.final_text])
    wb.save(report_path)


def _refresh_totals(batch: MessageBatch, *, mark_finished: bool = False) -> None:
    counts = dict(batch.recipients.order_by().values("status").annotate(c=Count("id")).values_list("status", "c"))
    batch.total_rows = sum(counts.values())
    batch.total_sent = counts.get(MessageRecipient.Status.SENT, 0)
    batch.total_failed = counts.get(MessageRecipient.Status.FAILED, 0)
    batch.total_invalid = counts.get(MessageRecipient.Status.INVALID_PHONE, 0)
    batch.total_duplicate = counts.get(MessageRecipient.Status.DUPLICATE, 0)
    batch.total_not_bale_user = counts.get(MessageRecipient.Status.NOT_BALE_USER, 0)
    batch.total_rate_limited = counts.get(MessageRecipient.Status.RATE_LIMITED, 0)
    batch.total_payment_required = counts.get(MessageRecipient.Status.PAYMENT_REQUIRED, 0)
    batch.total_config_error = counts.get(MessageRecipient.Status.CONFIG_ERROR, 0)
    update_fields = [
        "total_rows",
        "total_sent",
        "total_failed",
        "total_invalid",
        "total_duplicate",
        "total_not_bale_user",
        "total_rate_limited",
        "total_payment_required",
        "total_config_error",
    ]
    if mark_finished:
        batch.finished_at = timezone.now()
        update_fields.append("finished_at")
    batch.save(update_fields=update_fields)


def _is_cancel_requested(batch: MessageBatch) -> bool:
    batch.refresh_from_db(fields=["cancel_requested"])
    return batch.cancel_requested


def _sleep_until_cancel_or_timeout(batch: MessageBatch, seconds: float) -> bool:
    """Sleep in small chunks and return True if cancellation was requested."""
    if seconds <= 0:
        return _is_cancel_requested(batch)

    remaining = seconds
    while remaining > 0:
        if _is_cancel_requested(batch):
            return True
        chunk = min(0.2, remaining)
        sleep(chunk)
        remaining -= chunk
    return _is_cancel_requested(batch)


def process_excel_batch(*, batch: MessageBatch, file_path: str, sleep_seconds: float | None = None, sheet_name: str | None = None, skip_duplicates: bool = True, report_path: str | None = None) -> MessageBatch:
    batch.refresh_from_db(fields=["cancel_requested"])
    if batch.cancel_requested:
        batch.status = MessageBatch.Status.CANCELLED
        batch.error_message = "پردازش قبل از شروع با درخواست کاربر متوقف شد."
        batch.finished_at = timezone.now()
        batch.save(update_fields=["status", "error_message", "finished_at"])
        return batch

    batch.status = MessageBatch.Status.RUNNING
    batch.error_message = ""
    batch.finished_at = None
    batch.save(update_fields=["status", "error_message", "finished_at"])

    try:
        recipients = read_excel_recipients(file_path, sheet_name=sheet_name)
        if batch.limit:
            recipients = recipients[: batch.limit]

        seen: set[str] = set()
        client = BaleSafirClient()
        delay = settings.BALE_DEFAULT_SLEEP_SECONDS if sleep_seconds is None else sleep_seconds
        pending_bulk_recipients: list[MessageRecipient] = []

        def queue_bulk_recipient(obj: MessageRecipient) -> None:
            pending_bulk_recipients.append(obj)
            if len(pending_bulk_recipients) >= 1000:
                MessageRecipient.objects.bulk_create(pending_bulk_recipients, batch_size=1000)
                pending_bulk_recipients.clear()

        def flush_bulk_recipients() -> None:
            if pending_bulk_recipients:
                MessageRecipient.objects.bulk_create(pending_bulk_recipients, batch_size=1000)
                pending_bulk_recipients.clear()

        cancelled = False
        for item in recipients:
            if _is_cancel_requested(batch):
                cancelled = True
                break

            final_text = render_message(batch.message_template, item)
            obj = MessageRecipient(
                batch=batch,
                row_number=item.row_number,
                first_name=item.first_name,
                last_name=item.last_name,
                full_name=item.full_name,
                raw_phone=item.raw_phone,
                normalized_phone=item.normalized_phone or "",
                final_text=final_text,
            )

            if not item.normalized_phone:
                obj.status = MessageRecipient.Status.INVALID_PHONE
                obj.error_message = "شماره موبایل معتبر نیست."
                queue_bulk_recipient(obj)
                continue

            if skip_duplicates and item.normalized_phone in seen:
                obj.status = MessageRecipient.Status.DUPLICATE
                obj.error_message = "این شماره در همین فایل تکراری است."
                queue_bulk_recipient(obj)
                continue
            seen.add(item.normalized_phone)

            request_id = f"bale-{batch.id}-{item.row_number}-{uuid4().hex[:8]}"
            obj.request_id = request_id

            if batch.dry_run:
                obj.status = MessageRecipient.Status.DRY_RUN
                obj.api_message = "dry-run: ارسال واقعی انجام نشد."
                queue_bulk_recipient(obj)
                continue

            obj.save()
            result = client.send(
                phone_number=item.normalized_phone,
                text=final_text,
                request_id=request_id,
                button_text=batch.button_text,
                button_url=batch.button_url,
            )
            obj.status = result.get("status", MessageRecipient.Status.FAILED)
            obj.http_status = result.get("http_status")
            obj.api_code = result.get("api_code", "")
            obj.api_message = result.get("api_message", "")
            obj.raw_response = result.get("raw_response", "")
            obj.error_message = result.get("error", "")
            obj.sent_at = timezone.now() if obj.status == MessageRecipient.Status.SENT else None
            obj.save(
                update_fields=[
                    "status",
                    "http_status",
                    "api_code",
                    "api_message",
                    "raw_response",
                    "error_message",
                    "sent_at",
                ]
            )
            if _is_cancel_requested(batch):
                cancelled = True
                break
            if delay and _sleep_until_cancel_or_timeout(batch, delay):
                cancelled = True
                break

        flush_bulk_recipients()
        if not report_path:
            reports_dir = Path(settings.BASE_DIR) / "reports"
            report_path = str(reports_dir / f"bale_report_batch_{batch.id}.xlsx")
        batch.report_path = report_path
        batch.save(update_fields=["report_path"])
        write_report(batch, Path(report_path))
        _refresh_totals(batch)
        if cancelled or batch.cancel_requested:
            batch.status = MessageBatch.Status.CANCELLED
            batch.error_message = "پردازش با درخواست کاربر متوقف شد. گزارش شامل ردیف‌هایی است که تا قبل از توقف پردازش شده‌اند."
            batch.finished_at = timezone.now()
            batch.save(update_fields=["status", "error_message", "finished_at"])
        elif not batch.dry_run and (batch.total_failed or batch.total_config_error):
            batch.status = MessageBatch.Status.FAILED
            batch.error_message = "پردازش کامل شد، اما بخشی از ارسال‌ها به دلیل خطای API یا تنظیمات ناموفق بود. جزئیات ردیف‌ها را بررسی کن."
            batch.finished_at = timezone.now()
            batch.save(update_fields=["status", "error_message", "finished_at"])
        else:
            batch.status = MessageBatch.Status.FINISHED
            batch.finished_at = timezone.now()
            batch.save(update_fields=["status", "finished_at"])
    except Exception as exc:
        _refresh_totals(batch)
        batch.status = MessageBatch.Status.FAILED
        batch.error_message = f"خطا در پردازش batch: {exc}"
        batch.finished_at = timezone.now()
        batch.save(update_fields=["status", "error_message", "finished_at"])
        raise

    return batch


def send_single_recipient_test(*, first_name: str, last_name: str, phone: str, message_template: str, button_text: str | None = None, button_url: str | None = None) -> MessageBatch:
    item = ExcelRecipient(
        row_number=1,
        first_name=first_name.strip(),
        last_name=last_name.strip(),
        raw_phone=phone.strip(),
        normalized_phone=normalize_iran_mobile(phone),
    )
    final_text = render_message(message_template, item)
    batch = MessageBatch.objects.create(
        source_file_name="ارسال تست تک‌نفره",
        message_template=message_template,
        button_text=button_text or "",
        button_url=button_url or "",
        dry_run=False,
        limit=1,
        status=MessageBatch.Status.RUNNING,
    )
    recipient = MessageRecipient.objects.create(
        batch=batch,
        row_number=item.row_number,
        first_name=item.first_name,
        last_name=item.last_name,
        full_name=item.full_name,
        raw_phone=item.raw_phone,
        normalized_phone=item.normalized_phone or "",
        final_text=final_text,
    )

    if not item.normalized_phone:
        recipient.status = MessageRecipient.Status.INVALID_PHONE
        recipient.error_message = "شماره موبایل معتبر نیست."
        recipient.save(update_fields=["status", "error_message"])
    else:
        recipient.request_id = f"bale-test-{batch.id}-{uuid4().hex[:8]}"
        result = BaleSafirClient().send(
            phone_number=item.normalized_phone,
            text=final_text,
            request_id=recipient.request_id,
            button_text=button_text,
            button_url=button_url,
        )
        recipient.status = result.get("status", MessageRecipient.Status.FAILED)
        recipient.http_status = result.get("http_status")
        recipient.api_code = result.get("api_code", "")
        recipient.api_message = result.get("api_message", "")
        recipient.raw_response = result.get("raw_response", "")
        recipient.error_message = result.get("error", "")
        recipient.sent_at = timezone.now() if recipient.status == MessageRecipient.Status.SENT else None
        recipient.save(
            update_fields=[
                "request_id",
                "status",
                "http_status",
                "api_code",
                "api_message",
                "raw_response",
                "error_message",
                "sent_at",
            ]
        )

    reports_dir = Path(settings.BASE_DIR) / "reports"
    batch.report_path = str(reports_dir / f"bale_report_single_test_{batch.id}.xlsx")
    batch.save(update_fields=["report_path"])
    write_report(batch, Path(batch.report_path))
    _refresh_totals(batch, mark_finished=True)
    if recipient.status == MessageRecipient.Status.SENT:
        batch.status = MessageBatch.Status.FINISHED
        batch.error_message = ""
        batch.save(update_fields=["status", "error_message"])
    else:
        batch.status = MessageBatch.Status.FAILED
        batch.error_message = "ارسال تست موفق نبود. جزئیات خطا را در گزارش همین ارسال ببین."
        batch.save(update_fields=["status", "error_message"])
    return batch


def run_excel_batch(*, file_path: str, message_template: str, dry_run: bool = True, limit: int | None = None, sleep_seconds: float | None = None, sheet_name: str | None = None, button_text: str | None = None, button_url: str | None = None, skip_duplicates: bool = True, report_path: str | None = None) -> MessageBatch:
    batch = MessageBatch.objects.create(
        source_file_name=Path(file_path).name,
        message_template=message_template,
        button_text=button_text or "",
        button_url=button_url or "",
        dry_run=dry_run,
        limit=limit,
    )
    return process_excel_batch(
        batch=batch,
        file_path=file_path,
        sleep_seconds=sleep_seconds,
        sheet_name=sheet_name,
        skip_duplicates=skip_duplicates,
        report_path=report_path,
    )
